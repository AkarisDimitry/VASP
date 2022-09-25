# ------------------------------------- #
# ---  XML XML XML XML XML XML XML  --- #
# ------------------------------------- #
print('Python 3.X - https://www.python.org/download/releases/3.0/')
#####################################
# 		 Load python libraries 		#
#####################################

# === === === import libraries === === === #
# *** warning supresion
import warnings
warnings.filterwarnings("ignore")

# *** warning supresion
import warnings; warnings.filterwarnings("ignore")

# *** python libraries
try:
	import itertools, operator, logging, time, copy, pickle, datetime, os, sys, argparse
	#os.chmod('.', 777)
except:  print('WARNING :: XML.import_libraries() :: can not import itertools, operator, logging, time, copy, pickle or os')

# *** numpy libraries
try:
	import numpy as np
except: print('WARNING :: XML.import_libraries() :: can not import numpy ')


# XML #
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Border, Side, Color, PatternFill, NamedStyle, Alignment
from openpyxl.chart import BarChart, Series, Reference

# PANDAs #
import pandas as pd

# PLOT #
# *** matplotlib libraries 
try:
	# seaborn #
	import seaborn as sns

	# matplotlib #
	import matplotlib.pyplot as plt
	import matplotlib.axes as ax
	import matplotlib.patches as patches

	# ASE #
	from ase import Atoms
	from ase.visualize import view
	from ase.visualize.plot import plot_atoms
	from ase.lattice.cubic import FaceCenteredCubic
except:	print('WARNING :: XML.import_libraries() :: can not import matplotlib ')


try:	from src import POSCAR
except:	
	try: import POSCAR as POSCAR
	except: print('WARNING :: XML.import_libraries() :: can not import POSCAR ')

try:	from src import DOSCAR
except:	
	try: import DOSCAR as DOSCAR
	except: print('WARNING :: XML.import_libraries() :: can not import DOSCAR ')

try:	from src import OUTCAR
except:	
	try: import OUTCAR as OUTCAR
	except: print('WARNING :: XML.import_libraries() :: can not import OUTCAR ')

try:	from src import OSZICAR
except:	
	try: import OSZICAR as OSZICAR
	except: print('WARNING :: XML.import_libraries() :: can not import OSZICAR ')

try:	from src import Data
except:	
	try: import Data as Data
	except: print('WARNING :: XML.import_libraries() :: can not import Data ')

try:	from src import System
except:	
	try: import System as System
	except: print('WARNING :: XML.import_libraries() :: can not import System ')

try:	from src import ORR
except:	
	try: import ORR as ORR
	except: print('WARNING :: XML.import_libraries() :: can not import ORR ')

try:	from src import Set
except:	
	try: import Set as Set
	except: print('WARNING :: XML.import_libraries() :: can not import ORR ')
from Set import *

try:
	from scipy.signal import savgol_filter

	# NON linear models #
	from sklearn.neural_network import MLPRegressor
	from sklearn.datasets import make_regression
	from sklearn.model_selection import train_test_split
	from sklearn.model_selection import ShuffleSplit # or StratifiedShuffleSplit
	# linear models #
	from sklearn.decomposition import PCA
	from sklearn.preprocessing import StandardScaler
	from sklearn import linear_model
	from sklearn.model_selection import cross_val_predict
	from sklearn.metrics import mean_squared_error, r2_score

except:	print('ERROR :: DATA.import_libraries() :: can not import sklearn')

class XML(object):
	def __init__(self, 	name=None, sheet=None, 
						DATA=None, path=None, ):
		self.name  = name
		self.sheet = []
		self.path  = path if not path is None else '.'

		# == DATA == #
		self.set = DATA
		self.set_filename = None

		self.ORR_array = None
		self.ORR_df    = None
		self.ORR_relevant = [	'G1_ORR','G2_ORR','G3_ORR','G4_ORR','limiting_step_ORR',
								'Eabs_O','Gabs_O','Eabs_OH','Gabs_OH','Eabs_OOH','Gabs_OOH',
								'overpotencial_ORR_4e']
		self.ORR_activated = True

		self.Magnetization_array = None
		self.Magnetization_df    = None

		self.Distance_array = None
		self.Distance_df    = None
		

		self.color_names = ['yellow1' , 'yellow2', 'yellow3',
							'red1'    , 'red2'   , 'red3'   , 	
							'green1'  , 'green2' , 'green3' ,	
							'blue1'   , 'blue2'  , 'blue3'  , 'blue4', 
							'orange1' , 'orange2', 'orange3', 'orange4',
							'lightred', 'grey'   , ]

		self.color_HEX = [  'FFEB99', 'faf25f', 'FFEB99',
							'FCB1B1', 'ff5959', 'FCB1B1', 
							'B0F2C2', '73ff85', 'B0F2C2', 
							'A4C5C6', '88bbbd', 'A4C5C6', 'A4C5C6',
							'F5CDAA', 'edbd93', 'F5CDAA', 'F5CDAA',	
							'F5DCF9', 'c2c2c2', ]

		self.color = { name: PatternFill(start_color=color, end_color=color, fill_type='solid') for name, color in zip(self.color_names, self.color_HEX) }

		self.WB = openpyxl.Workbook()

	def save(self, filename='summary.xlsx'):	self.WB.save('{}'.format(filename) )

	def set_load(self, filename, save=True):
		# load pickle file {filename} # 
		set_temp = Set()
		set_temp.load_data(filename)
 
		if save: self.set = set_temp

		return set_temp

	def make_path(self, path):
		if not os.path.isdir(path):  	 			
			os.makedirs(path)
			return True

		return False

	def creat_styles(self, WB=None):
		WB = WB if not WB is None else self.WB

		highlight = NamedStyle(name="highlight")
		highlight.font = Font(bold=True, size=20)
		bd = Side(style='thick', color="000000")
		highlight.border = Border(left=bd, top=bd, right=bd, bottom=bd)
		WB.add_named_style(highlight)

		return True

	def plot_bars(self, data, x, y, hue, 
					   Title='', path=None, name='plot.png',
					    ):
		path = path if not path is None else self.path
		if type(data) != type(None) and len(data) > 0 and not os.path.exists(f"{path}/{name}"):
			plt.clf(); 	plt.cla();  plt.close()
			ax = sns.barplot(
			    x=x, 
			    y=y, 
			    hue=hue, 
			    data=data, 
			    ci="sd", 
			    edgecolor="black",
			    errcolor="black",
			    errwidth=1.5,
			    capsize = 0.1,
			    alpha=0.5
			)
			sns.stripplot(
			    x=x, 
			    y=y, 
			    hue=hue, 
			    data=data, dodge=True, alpha=0.6, ax=ax
			)

			handles, labels = ax.get_legend_handles_labels()
			ax.legend(handles, labels, title=hue, bbox_to_anchor=(1, 1.02), loc='upper left')
			ax.set_xticklabels(ax.get_xticklabels(), rotation=45)
			ax.set_title(Title)
			fig = ax.get_figure()
			fig.savefig(f"{path}/{name}" , dpi=200, pad_inches=0.1, bbox_inches='tight', horizontalalignment='right') 

			return ax

		return False

	def plot_add(self, path, name, anchor, sheet, height=400, width=400):
		img = openpyxl.drawing.image.Image(f"{path}/{name}")
		img.anchor = anchor
		img.height = height # insert image height in pixels as float or int (e.g. 305.5)
		img.width  = width  # insert image width in pixels as float or int (e.g. 405.8)
		sheet.add_image(img)

		return True

	def summary(self, path=None):
		path = path if not path is None else self.path
		# fill the workbook with all possible data from DATA.py #

		# ====== Create folthers ====== #
		self.make_path(f'{path}/plots/geometrias/')
		self.make_path(f'{path}/plots/ORR/')
		self.make_path(f'{path}/plots/metanalisis/system')
		self.make_path(f'{path}/plots/metanalisis/Cell')
		self.make_path(f'{path}/plots/metanalisis/Functional')
		self.make_path(f'{path}/plots/metanalisis/magnetization')
		self.make_path(f'{path}/plots/metanalisis/magnetization/System')
		self.make_path(f'{path}/plots/metanalisis/magnetization/Functional/')
		self.make_path(f'{path}/plots/metanalisis/distance')
		self.make_path(f'{path}/plots/metanalisis/distance/System')
		self.make_path(f'{path}/plots/metanalisis/distance/Functional/')

		self.get_ORR_array()
		self.get_ORR_DF()

		self.get_Magnetization_array()
		self.get_Magnetization_DF()

		self.get_Distance_array()
		self.get_Distance_DF()

		self.add_magnetization_sheet()
		self.add_distance_sheet()
		self.add_ORR_sheet()

		self.summary_ORR()
		self.summary_set()

		self.save()


	def summary_set(self, set_full=None, path=None, plot=True, WB=None):
		set_full = set_full if not set_full is None else self.set.set
		path 	 = path 	if not path 	is None else self.path
		WB = WB if not WB is None else self.WB

		for set_n, (key_data, data) in enumerate(set_full.items()):
			self.summary_data(data=data, path=path, plot=plot, key_data=key_data, WB=WB ) 

					
	def summary_data(self, data, path=None, plot=True, key_data=None, WB=None):
		path 	 = path 	if not path 	is None else self.path
		WB = WB if not WB is None else self.WB

		self.make_path(f'{path}/plots/systemas/{key_data}')

		# == make new sheet == (with an acceptable name)
		sheet_name = key_data.replace('*','x') 
		sheet_name = sheet_name.replace('[','L') 
		sheet_name = sheet_name.replace(']','L') 

		ws    = WB.create_sheet(sheet_name) # insert at the end (default)
		sheet = WB.get_sheet_by_name(sheet_name)

		if self.ORR_activated:
			data.ORR()
			ORR = data.reaction('ORR')
			# == make plot == # 
			if not os.path.exists(f"{path}/plots/ORR/{key_data}.png"):
				ORR.plot(folder=f'{path}/plots/ORR', name=key_data)

			# == if exist add plot to sheet == #
			if os.path.exists(f"{path}/plots/ORR/{key_data}.png"):
				self.plot_add(	path  =f"{path}/plots/ORR", 
								name  =f"{key_data}.png", 
								anchor=f'{get_column_letter(1)}1', 
								sheet =sheet, height=400, width=400)		
			else:
				print(f'WARNNING :: Can not add {path}/plots/ORR/{key_data}.png')

		# == set some sheer variables == #
		#sheet.column_dimensions['A'].width = 30
		for system_n, (key_system, system) in enumerate(data.system.items()):
			self.summary_system(system=system, key_data=key_data, key_system=key_system, 
								path=path, plot=plot, ID=system_n, sheet=sheet)
			
	def summary_system(self, system, key_system=None,  key_data=None,
							 path=None, plot=True, ID=1, sheet=None):
		# ======================= PLOT ATOMS ======================= #
		try:
			if plot and not system.CONTCAR is None:	
				if not os.path.exists(f'{path}/plots/systemas/{key_data}/{key_system}.png'):
					system.CONTCAR.plot(save=True, path=f'{path}/plots/systemas/{key_data}/{key_system}.png')
				
				if os.path.exists(f'{path}/plots/systemas/{key_data}/{key_system}.png'):
					self.plot_add(	path  =f"{path}/plots/systemas/{key_data}", 
									name  =f"{key_system}.png", 
									anchor=f'{get_column_letter(1+18*ID)}20', 
									sheet =sheet, height=400, width=400)	

		except: print('')

		for load in system.loaded:
			# = Files name = #
			#sheet[f'D{line}'].value  = load
			#sheet[f'D{line}'].font   = Font(name="Tahoma", size=13, color="00339966", bold=True)
			#sheet[f'D{line}'].border = Border( right=double )			
			
			# = Files summary = #
			try:
				if 	 load == 'POSCAR': system_resume = system.POSCAR.resume()
				elif load == 'CONTCAR': system_resume = system.CONTCAR.resume()
				elif load == 'OSZICAR': system_resume = system.OSZICAR.resume()
				elif load == 'OUTCAR': 
					system_resume = system.OUTCAR.resume()

					for nx in range(3):
						sheet.cell(row=45+nx, column=4).value =  'cell'
						sheet.cell(row=45+nx, column=5).value = f'v{nx}'
						sheet.cell(row=45+nx, column=5).fill = self.color['grey']
						for ny in range(3):
							sheet.cell(row=45+nx, column=6+ny).value = '{:.3f}'.format(system_resume['dict']['cell'][nx][ny])
							sheet.cell(row=45+nx, column=5).fill = self.color['red1']
							
					for cli, cl in enumerate(['','N','ID','X','Y','Z','fX','fY','fZ',
												'chaS','chaP','chaD','chaT',
												'magS','magP','magD','magT']):
						sheet.cell(row=49, column=cli + 18*ID + 1).value = cl
						sheet.cell(row=49, column=cli + 18*ID + 1).fill = self.color['lightred']

					for i, n in enumerate(system_resume['dict']['atoms_names_list']):
						sheet.cell(row=50+i, column=1 + 18*ID).value = f'atom'	
						sheet.column_dimensions[get_column_letter(1+18*ID)].width = 4.5
						sheet.cell(row=50+i, column=2 + 18*ID).value = f'{i}'	
						sheet.column_dimensions[get_column_letter(2+18*ID)].width = 4

						sheet.cell(row=50+i, column=3 + 18*ID).value = system_resume['dict']['atoms_names_list'][i]
						sheet.column_dimensions[get_column_letter(3+18*ID)].width = 3

						for cl in range(4,18): 
							sheet.column_dimensions[get_column_letter(cl+18*ID)].width = 7
							sheet.cell(row=50+i, column=cl + 18*ID).alignment = Alignment(horizontal='right')  # right justification

						for cl in range(6):	sheet.cell(row=50+i, column=4 +cl+18*ID).value = '{:.3f}'.format(system_resume['dict']['total_force'][i][cl])
						for cl in range(4):	sheet.cell(row=50+i, column=10+cl+18*ID).value = '{:.3f}'.format(system_resume['dict']['total_charge'][i][cl])
						for cl in range(4):	sheet.cell(row=50+i, column=14+cl+18*ID).value = '{:.3f}'.format(system_resume['dict']['magnetization'][i][cl])
 	
						colors0 = [ self.color[cn] for cn in [	'yellow1' , 'yellow1', 'yellow1',
																'red1'    , 'red1'   , 'red1'   , 
															 	'green1'  , 'green1' , 'green1' ,
															 	'blue1'   , 'blue1'  , 'blue1'  , 'blue1', 
															 	'orange1' , 'orange1', 'orange1', 'orange1' ]  ]

						colors1 = [ self.color[cn] for cn in [	'yellow2' , 'yellow2', 'yellow2',
																'red2'    , 'red2'   , 'red2'   , 
															 	'green2'  , 'green2' , 'green2' ,
															 	'blue2'   , 'blue2'  , 'blue2'  , 'blue2', 
															 	'orange2' , 'orange2', 'orange2', 'orange2' ]  ]
						for cl in range(17): 
							if i%2==0: sheet.cell(row=50+i, column=1 +cl+18*ID).fill = colors0[cl]
							if i%2==1: sheet.cell(row=50+i, column=1 +cl+18*ID).fill = colors1[cl]

					cl =  55+i
					for atom_id_0i, (atom_id_0, atom_N) in enumerate(system.CONTCAR.relevant_distances().items()):
 						# 'O': {0: {'distances': array([1.68028891]), 'index': array([65]), 'names': ['Co']}}
						for atom_N_0i, (atom_N_0, data) in enumerate(atom_N.items()):
							# {0: {'distances': array([1.68028891]), 'index': array([65]), 'names': ['Co']}}
							for atom_id_1i, atom_id_1 in enumerate(data['names']):
								# 'names': ['Co']
								sheet.cell(row=cl, column=6  + 18*ID).fill  = self.color['red1']
								sheet.cell(row=cl, column=6  + 18*ID).value = f'{atom_id_0}'								# ID 0 #

								sheet.cell(row=cl, column=7  + 18*ID).value = f'{atom_N_0}'									# atom number 0 #
								sheet.cell(row=cl, column=7  + 18*ID).fill  = self.color['grey']

								sheet.cell(row=cl, column=8  + 18*ID).value = f'{atom_id_1}'								# ID 1 #
								sheet.cell(row=cl, column=8  + 18*ID).fill  = self.color['red1']
								
								sheet.cell(row=cl, column=9  + 18*ID).value = '{}'.format(data['index'][atom_id_1i])		# atom number 1 #
								sheet.cell(row=cl, column=9  + 18*ID).fill  = self.color['grey']

								sheet.cell(row=cl, column=10 + 18*ID).value = '{:.3f}'.format(data['distances'][atom_id_1i])	# Distance #
								sheet.cell(row=cl, column=10  + 18*ID).fill  = self.color['blue1']
								cl += 1

				else: system_resume= {'dict':{}}

			except:	print('WARNING :: Set.summary() :: can NOT plot summarise POSCAR/CONTCAR/OSZICAR/OUTCAR')


	# === === === ORR summary === === === # 
	def summary_ORR(self, ORR_df=None, path=None, WB=None):
		ORR_df = ORR_df if not ORR_df is None else self.ORR_df 
		WB     = WB     if not WB     is None else self.WB 
		path   = path   if not path   is None else self.path 

		# ORR_DF : columns : ORR_relevant+['FullID', 'System', 'Cell', 'Functional']
		for orr_r in self.ORR_relevant:	
			ws       = WB.create_sheet(orr_r) # insert at the end (default)
			ORRsheet = WB.get_sheet_by_name(orr_r)

			# ============= PLOT system ( system name ) ============= #
			self.make_path(f'{path}/plots/metanalisis/system/{orr_r}')
			line = 0
			for system in pd.unique(ORR_df['System']):
				if not os.path.exists(f"{path}/plots/metanalisis/system/{orr_r}/bars_plot_{system}.png"):
					self.plot_bars(ORR_df[ ORR_df['System']== system ], x="Functional", y=orr_r, hue="Cell", 
									Title=f'System: {system} | ORR {orr_r} ', 
									path =f"{path}/plots/metanalisis/system/{orr_r}", 
									name =f"bars_plot_{system}.png")
				
				self.plot_add(	path  =f"{path}/plots/metanalisis/system/{orr_r}", 
								name  =f"bars_plot_{system}.png", 
								anchor=f'{get_column_letter((line%3)*5+1 )}{int(line/3)*20+1}', 
								sheet =ORRsheet, height=400, width=400)		
				line += 1

			# ============= PLOT name3 ( system functional ) ============= #
			self.make_path(f"{path}/plots/metanalisis/Functional/{orr_r}")
			line += 10
			for functional in pd.unique(ORR_df['Functional']):
				if not os.path.exists(f"{path}/plots/metanalisis/Functional/{orr_r}/bars_plot_{functional}.png"):
					self.plot_bars(ORR_df[ ORR_df['Functional']== functional ], x="System", y=orr_r, hue="Cell", 
									Title=f'Functional: {functional} | ORR {orr_r} ', 
									path =f"{path}/plots/metanalisis/Functional/{orr_r}", 
									name =f"bars_plot_{functional}.png")
				
				self.plot_add(	path  =f"{path}/plots/metanalisis/Functional/{orr_r}", 
								name  =f"bars_plot_{functional}.png", 
								anchor=f'{get_column_letter((line%3)*5+1 )}{int(line/3)*20+1}', 
								sheet =ORRsheet, height=400, width=400)		
				line += 1

		return True

	# === === === ORR getter === === === # 
	def get_ORR_array(self, save=True, plot=True, path=None):
		path = path if not path is None else self.path

		# set conteiners #
		ORR_relevant = self.ORR_relevant
		ORR_array   = []
		for set_n, (key_data, data) in enumerate(self.set.set.items()):
			# -- generate OBJ and alalize the ORR reaction -- #
			data.ORR()
			ORR = data.reaction('ORR')

			# -- generate plot -- #
			if plot and not os.path.exists(f"{path}/plots/ORR/{key_data}.png"):
				ORR.plot(folder=f'{path}/plots/ORR/.', name=key_data)

			# -- store DATA in vector-- #
			if type(ORR.ORR) == dict:	 # store ORR info of all systems #
				vector_labels  = [key_data, key_data.split('_')[0], key_data.split('_')[1], '_'.join(key_data.replace('__','_').split('_')[2:]) ] 
				vector_numeric = [ORR.ORR[label] for label in ORR_relevant]

			# == ORR_array data append == #
			ORR_array.append(vector_numeric+vector_labels)

		ORR_array = np.array(ORR_array) # store ORR info of all systems #

		if save:	self.ORR_array = ORR_array

		return ORR_array

	def get_ORR_DF(self, ORR_array=None, ORR_relevant=None, save=True):
		ORR_array = ORR_array if not ORR_array is None else self.ORR_array
		ORR_relevant = ORR_relevant if not ORR_relevant == None else self.ORR_relevant

		# -- make ORR dataframe -- #
		ORR_df  	= pd.DataFrame(data=ORR_array,  columns=ORR_relevant+['FullID', 'System', 'Cell', 'Functional']) # store ORR info of all systems #
		ORR_df[ORR_relevant] = ORR_df[ORR_relevant].apply(pd.to_numeric)
		if save:	self.ORR_df = ORR_df
		return ORR_df
	
	# === === === Magnetization getter === === === # 
	def get_Magnetization_array(self, save=True, plot=True):
		# set conteiners #
		Magnetization_array = []
		
		# iterate each data conteiner #
		for set_n, (key_data, data) in enumerate(self.set.set.items()):
			vector_labels  = [key_data, key_data.split('_')[0], key_data.split('_')[1], '_'.join(key_data.replace('__','_').split('_')[2:]) ] 

			# iterate each system  #
			for system_n, (key_system, system) in enumerate(data.system.items()):
				key_subtype = key_system.split('_')[-1] 
				try:
					if key_subtype in ['*', '*O', '*OH', '*OOH']:
						# -- store DATA in vector-- #
						atoms_names =  np.array(system.OUTCAR.atoms_names_list)[system.CONTCAR.relevant_atoms_mask()] 
						atoms_magnetization = np.abs(system.OUTCAR.resume()['dict']['magnetization'][system.CONTCAR.relevant_atoms_mask()][:,-1])
						Magnetization_array += [ vector_labels+[key_subtype, atoms_names[atom_i], atom_i, atoms_magnetization[atom_i]] for atom_i, atom_name in enumerate(atoms_names)  ]	
				except: 
					print('ERROR :: XML.get_Magnetization_array() :: Can not Extrar magnetization data from dataset')
		
		Magnetization_array = np.array(Magnetization_array)

		if save: self.Magnetization_array = Magnetization_array

		return Magnetization_array

	def get_Magnetization_DF(self, Magnetization_array=None, save=True):
		Magnetization_array = Magnetization_array if not Magnetization_array is None else self.Magnetization_array

		# -- make MARNETIZATION dataframe -- #
		Magnetization_array = np.array(Magnetization_array)
		Magnetization_df  	= pd.DataFrame(data=Magnetization_array,  columns= ['FullID', 'Name', 'cell', 'Functional']+['System', 'Atom', 'number', 'Magnetization'] ) # store ORR info of all systems #
		Magnetization_df['Magnetization'] = Magnetization_df['Magnetization'].apply(pd.to_numeric)
		
		if save: self.Magnetization_df = Magnetization_df
		
		return Magnetization_df

	# === === === Distance getter === === === # 	
	def get_Distance_array(self, save=True ):
		# set conteiners #
		Distance_array = []

		# iterate each data conteiner #
		for set_n, (key_data, data) in enumerate(self.set.set.items()):
			vector_labels  = [key_data, key_data.split('_')[0], key_data.split('_')[1], '_'.join(key_data.replace('__','_').split('_')[2:]) ] 
			
			# iterate each system  #
			for system_n, (key_system, system) in enumerate(data.system.items()):
				key_subtype = key_system.split('_')[-1] 
				try:
					if key_subtype in ['*', '*O', '*OH', '*OOH']:
						# correct CONTCAR ID with OUTCAR data # 
						system.CONTCAR.atoms_names_list = system.OUTCAR.atoms_names_list
						system.CONTCAR.atoms_names_ID 	= system.OUTCAR.atoms_names_ID
						#system.CONTCAR.atoms_names_full = system.OUTCAR.atoms_names_full

						# -- store DATA in vector-- #
						Distance_vec = [[ vector_labels+[key_subtype, atom1, atom2, prop['distances'][pos]] for pos, atom2 in enumerate(prop['names']) ] for atom1, atom1N in system.CONTCAR.relevant_distances().items()  for num, prop in atom1N.items()  ] 
						for n in Distance_vec:
							Distance_array += n
				except: pass

		Distance_array = np.array(Distance_array)
		if save: self.Distance_array = Distance_array

		return Distance_array

	def get_Distance_DF(self, Distance_array=None, save=True):
		# set conteiners #
		Distance_array = Distance_array if not Distance_array is None else self.Distance_array
		
		# -- make DISTANCE dataframe -- #
		Distance_array = np.array(Distance_array)
		Distance_df  	 = pd.DataFrame(data=Distance_array,  columns= ['FullID', 'Name', 'cell', 'Functional']+['System', 'Atom1', 'Atom2', 'Distance'] ) # store ORR info of all systems #
		Distance_df['Distance'] = Distance_df['Distance'].apply(pd.to_numeric)

		if save: self.Distance_df = Distance_df

		return Distance_df
		
	# === === === MAKE SHEET magnetization === === === # 
	def add_magnetization_sheet(self, Magnetization_df=None, path=None, WB=None):
		path 			 = path 			if not path 			is None else self.path
		Magnetization_df = Magnetization_df if not Magnetization_df is None else self.Magnetization_df
		WB     			 = WB    			if not WB     			is None else self.WB 

		# == Magnetization Per ATOM/System/FUNCTIONAL == #  # == Magnetization Per ATOM/System/FUNCTIONAL == #  # == Magnetization Per ATOM/System/FUNCTIONAL == # 
		# -- Magnetization vs Functional -- #
		line = -2
		ws       = WB.create_sheet(f'Mag_functional') 
		MAGsheet = WB.get_sheet_by_name(f'Mag_functional')

		# -- FILTER 1 -- #
		data_0  = Magnetization_df[Magnetization_df['Magnetization'] < 10]
		for name_i in pd.unique(Magnetization_df['Name']):
			line += 1
			# -- FILTER 2 -- #
			data_1  = data_0[data_0['Name'] == name_i]

			for atom_i in pd.unique(Magnetization_df['Atom']):
				line += 1
				# -- FILTER 3 -- #
				data_2  = data_1[data_1['Atom'] == atom_i]

				for i, sys_i in enumerate(pd.unique(Magnetization_df['System'])):
					# == FILTER 4 == # 
					data_3  = data_2[data_2['System'] == sys_i]

					if len(data_3) > 0:			
						self.plot_bars(data_3, x="Functional", y="Magnetization", hue="Atom", 
										Title=f'Systems {sys_i} | Atom Type {atom_i} |  {name_i} ', 
										path =f"{path}/plots/metanalisis/magnetization/Functional", 
										name =f"bars_plot_{sys_i}_{atom_i}_{name_i}.png")

						self.plot_add(	path  =f"{path}/plots/metanalisis/magnetization/Functional", 
										name  =f"bars_plot_{sys_i}_{atom_i}_{name_i}.png", 
										anchor=f'{get_column_letter(i*5+1)}{line*20+1}', 
										sheet =MAGsheet, height=400, width=400)						
				
		# == Magnetization Per ATOM/System == #	# == Magnetization Per ATOM/System == # # == Magnetization Per ATOM/System == #		
		# Magnetization vs System #	
		line = -1
		ws       = WB.create_sheet(f'Mag_sys') # insert at the end (default)
		MAGsheet = WB.get_sheet_by_name(f'Mag_sys')

		# == FILTER 1B == # 
		data_0  = Magnetization_df[Magnetization_df['Magnetization'] < 10]
		for atom_i in pd.unique(Magnetization_df['Atom']):
			line += 1
			# == FILTER 2B == # 
			data_1  = data_0[data_0['Atom'] == atom_i]

			for i, sys_i in enumerate(pd.unique(Magnetization_df['System'])):
				# == FILTER 3B == # 
				data_2  = data_1[data_1['System'] == sys_i]

				if len(data_2) > 0:			
					self.plot_bars(data_2, x="Name", y="Magnetization", hue="Atom", 
									Title=f'Atom Type {atom_i} |  {name_i} ', 
									path =f"{path}/plots/metanalisis/magnetization/System", 
									name =f"bars_plot_{sys_i}_{atom_i}.png")

					self.plot_add(	path  =f"{path}/plots/metanalisis/magnetization/System", 
									name  =f"bars_plot_{sys_i}_{atom_i}.png", 
									anchor=f'{get_column_letter(i*5+1)}{line*20+1}', 
									sheet =MAGsheet, height=400, width=400)						
			
	# === === === MAKE SHEET distance === === === # 
	def add_distance_sheet(self, Distance_df=None,  path=None, WB=None):
		path 		= path        	if not path        is None else self.path
		Distance_df = Distance_df 	if not Distance_df is None else self.Distance_df
		WB 			= WB 		  	if not WB      	   is None else self.WB 
		# == Distance Per ATOM/System/FUNCTIONAL == #  # == Distance Per ATOM/System/FUNCTIONAL == #  # == Distance Per ATOM/System/FUNCTIONAL == # 

		# -- Distance vs Functional -- #
		line = -2
		ws        = WB.create_sheet(f'Dist_functional') 
		DISTsheet = WB.get_sheet_by_name(f'Dist_functional')
	
		# == FILTER 1A == # 
		data_0  = Distance_df[Distance_df['Distance'] < 10]
		for name_i in pd.unique(Distance_df['Name']):
			line += 1
			# == FILTER 2A == # 
			data_1  = data_0[data_0['Name'] == name_i]

			for atom_i in pd.unique(Distance_df['Atom1']):
				line += 1
				# == FILTER 3A == # 
				data_2  = data_1[data_1['Atom1'] == atom_i]

				for i, sys_i in enumerate(pd.unique(Distance_df['System'])):
					# == FILTER 2A == # 
					data_3  = data_2[data_2['System'] == sys_i]

					if len(data_3) > 0:			
						self.plot_bars(data_3, x="Functional", y="Distance", hue="Atom2", 
										Title=f'Distance in Systems {name_i}+{sys_i} | From:{atom_i} to...  ', 
										path =f"{path}/plots/metanalisis/distance/Functional", 
										name =f"bars_plot_{sys_i}_{atom_i}_{name_i}.png")

						self.plot_add(	path  =f"{path}/plots/metanalisis/distance/Functional", 
										name  =f"bars_plot_{sys_i}_{atom_i}_{name_i}.png", 
										anchor=f'{get_column_letter(i*5+1)}{line*20+1}', 
										sheet =DISTsheet, height=400, width=400)						
				
		# == Distance Per ATOM/System == #	# == Distance Per ATOM/System == # # == Distance Per ATOM/System == #		
		# Distance vs System #	
		line = -1
		ws        = WB.create_sheet(f'Dist_sys') # insert at the end (default)
		DISTsheet = WB.get_sheet_by_name(f'Dist_sys')
	
		# == FILTER 1A == # 
		data_0  = Distance_df[Distance_df['Distance'] < 10]
		for atom_i in pd.unique(Distance_df['Atom1']):
			line += 1
			# == FILTER 2A == # 
			data_1  = data_0[data_0['Atom1'] == atom_i]

			for i, sys_i in enumerate(pd.unique(Distance_df['System'])):
				# == FILTER 3A == # 
				data_2  = data_1[data_1['System'] == sys_i]

				if len(data_2) > 0:			
					self.plot_bars(data_2, x="Name", y="Distance", hue="Atom2", 
									Title=f'Atom Type {atom_i} |  {name_i} ', 
									path =f"{path}/plots/metanalisis/distance/System", 
									name =f"bars_plot_{sys_i}_{atom_i}.png")

					self.plot_add(	path  =f"{path}/plots/metanalisis/distance/System", 
									name  =f"bars_plot_{sys_i}_{atom_i}.png", 
									anchor=f'{get_column_letter(i*5+1)}{line*20+1}', 
									sheet =DISTsheet, height=400, width=400)	

	# === === === MAKE ORR distance === === === # 
	def add_ORR_sheet(self, ORR_df=None,  set_full=None, path=None, WB=None, plot=True):
		path 		= path        	if not path        is None else self.path
		ORR_df 	 	= ORR_df 		if not ORR_df 	   is None else self.ORR_df
		set_full 	= set_full 		if not set_full    is None else self.set.set
		WB 			= WB 		  	if not WB      	   is None else self.WB 
		# == ORR Per System == #  # == ORR Per System == #  # == ORR Per System == #  # == ORR Per System == #  # == ORR Per System == # 

		# === ORR TAB === #
		ws = WB.create_sheet('ORR') # insert at the end (default)
		ORRsheet = WB.get_sheet_by_name('ORR')
		ORRsheet.column_dimensions['A'].width = 30

		for i, label in enumerate(self.ORR_relevant):
			ORRsheet['{}{}'.format(get_column_letter(i+1), 2)].value = label
			ORRsheet['{}{}'.format(get_column_letter(i+1), 2)].font = Font(name="Tahoma", size=13, color="00339966", bold=True)
			#ORRsheet['{}{}'.format(get_column_letter(i+1), 2)].border = Border(top=double, bottom=double)

		ORRsheet = WB.get_sheet_by_name('ORR')

		for set_n, (key_data, data) in enumerate(set_full.items()):
			ORR = data.ORR()
			for i, label in enumerate(self.ORR_relevant):
				if type(ORR.ORR) != type(None):
					ORRsheet[f'{get_column_letter(i+2)}{set_n+3}'].value = ORR.ORR[label]

			# == make new sheet == (with an acceptable name)
			sheet_name = key_data.replace('*','x') 
			sheet_name = sheet_name.replace('[','L') 
			sheet_name = sheet_name.replace(']','L') 
			
			# == make LINK in ORR sheet == (with an acceptable name)
			#sheet['B1'].style = 'highlight'
			ORRsheet[f'A{set_n+3}'].hyperlink = './summary.xlsx'+f'#{sheet_name}!A1'
			ORRsheet[f'A{set_n+3}'].value = '#'+sheet_name
			ORRsheet[f'A{set_n+3}'].style = "Hyperlink"
			
			# === === === ORR === === === # # === === === ORR === === === # # === === === ORR === === === #

xml = XML(path='.')
xml.set_load(filename='/home/akaris/Documents/code/VASP/v4.6/files/dataset/CoFeTPyP/Catalisis/dataset_CoFeTPyPCoFe_catalysis.pkl')

xml.summary()

# poscar = POSCAR()










