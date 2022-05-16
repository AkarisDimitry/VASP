print('Python 3.X - https://www.python.org/download/releases/3.0/')
###################################
# Step 1 || Load python libraries #
###################################
# *** warning supresion
import warnings
warnings.filterwarnings("ignore")

# *** warning supresion
import warnings; warnings.filterwarnings("ignore")

# *** python libraries
try:
	import itertools, operator, logging, time, copy, pickle, datetime, os, sys, argparse
	#os.chmod('.', 777)
except:  print('WARNING :: DATA.import_libraries() :: can not import itertools, operator, logging, time, copy, pickle or os')

# *** numpy libraries
try:
	import numpy as np
except: print('WARNING :: DATA.import_libraries() :: can not import numpy ')

# *** matplotlib libraries 
try:
	import matplotlib.pyplot as plt
	import matplotlib.axes as ax
	import matplotlib.patches as patches
except:	print('WARNING :: Set.import_libraries() :: can not import matplotlib ')

try:	from src import POSCAR
except:	
	try: import POSCAR as POSCAR
	except: print('WARNING :: Set.import_libraries() :: can not import POSCAR ')

try:	from src import DOSCAR
except:	
	try: import DOSCAR as DOSCAR
	except: print('WARNING :: Set.import_libraries() :: can not import DOSCAR ')

try:	from src import OUTCAR
except:	
	try: import OUTCAR as OUTCAR
	except: print('WARNING :: Set.import_libraries() :: can not import OUTCAR ')

try:	from src import OSZICAR
except:	
	try: import OSZICAR as OSZICAR
	except: print('WARNING :: Set.import_libraries() :: can not import OSZICAR ')

try:	from src import Data
except:	
	try: import Data as Data
	except: print('WARNING :: Set.import_libraries() :: can not import Data ')

try:	from src import System
except:	
	try: import System as System
	except: print('WARNING :: Set.import_libraries() :: can not import System ')

try:	from src import ORR
except:	
	try: import ORR as ORR
	except: print('WARNING :: Set.import_libraries() :: can not import ORR ')


try:
	from scipy.signal import savgol_filter

	# NON linear models #
	from sklearn.neural_network import MLPRegressor
	from sklearn.datasets import make_regression
	from sklearn.model_selection import train_test_split
	from sklearn.model_selection import ShuffleSplit # or StratifiedShuffleSplit
	# linear models #
	from sklearn.decomposition import PCA
	from sklearn.preprocessing import StandardScaler
	from sklearn import linear_model
	from sklearn.model_selection import cross_val_predict
	from sklearn.metrics import mean_squared_error, r2_score

except:	print('ERROR :: DATA.import_libraries() :: can not import sklearn')

try:
	from ase import Atoms
	from ase.visualize import view
except:	print('ERROR :: DATA.import_libraries() :: can not import ase ')

class Set(object):
	def __init__(self, 	name=None, description=None, 
						path_list=None, name_list=None, ):
		self.name = name
		self.description = description

		self.path_list = path_list
		self.name_list = name_list

		self.set = {}

	def read_data(self, dataset=None, output=None, save=True, verbosity=True, log=True):
		'''# READ dataset #t
		dataset 	:: 		dict 	:: 	dictionnary  that conteins all path to data set elements
		save 		:: 		bool 	:: 	Indicates if its nesesary to save 
		verbosity	:: 		bool 	:: 	Indicates if its nesesary print information during the processes 
		'''
		v = verbosity
		output = output if type(output) == str else 'dataset_save_state' 
		dataset = dataset if type(dataset) == dict else self.dataset
		if not type(dataset) == dict: 	print('ERROR : code000 READ.read_data() : need dir/dataset ')

		with open(f'log.dat', 'a') as lod_file:
			lod_file.write(f'>> {datetime.datetime.now()} \t >> Set.read_data() \n')

		for set_name in dataset:
			if v: print('='*70)
			if v: print(f'>>  * {set_name} ')
			if v: print(f'    |  ')
			sub_system = dataset[set_name]
			data = Data.Data()

			for sys_name in sub_system: 
				#if v: print(f'    *=== {set_name}+{sys_name}')
				sys_path = sub_system[sys_name]
				data.load_system(	name='{}_{}'.format(set_name, sys_name), 
									path =sys_path, 
									files='all',
									v=False)
			data.ORR()
			data.summary( )

			self.set[str(set_name)] = data

		if save: self.save_set(filename=f'{output}.pkl', dataset=self.set, light=False)
		return True

	def save_set(self, filename=None, dataset=None, light=True):
		dataset = dataset if type(dataset) == dict else self.set


		#[for sys_name in dataset[set_name] for set_name in dataset ]
		if light:
			subsystem_order = { unique_name:index for index, unique_name in enumerate(list(set([sys_name for set_name in dataset for sys_name in dataset[set_name]  ])))}
			dataset_matrix = np.zeros( len(dataset), len(subsys_list) )
			for set_name in dataset:
				for sys_name in dataset[set_name]:
					pass # NOT IMPLEMENTED

		else:
			filename = filename if type(filename) else 'filename'
			filehandler = open(filename, 'wb') 
			pickle.dump(dataset, filehandler)

	def load_data(self, filename, verbosity=True):
		dataset = pickle.load( open( filename, "rb" ), encoding='latin1') # encoding='bytes'
		self.__dict__ = dataset.__dict__.copy() 
		print(f' >> Successfully read {filename}\n')
		return dataset

	def make_tabla(self, matrix, format='latex', col_names=[], fil_names=[], print=True, save=False ):
		lines = []
		lines.append( ' & '.join(col_names)+'\\\\' )
		for i, n in enumerate(matrix):
			lines.append( ' & '.join(n)+'\\\\' )

	def summary(self, feature={'ORR':True}, latex=False, xml=False, path=None):
		path = path if type(path) == str else '.' 

		for set_n, (key_data, data) in enumerate(self.set.items()):
			print( ' {1} System {0} {1}'.format(key_data, '*'*12) )
			data.summary( feature={'ORR':True} )
		
		# ------------------------------------- #
		# ---  XML XML XML XML XML XML XML  --- #
		# ------------------------------------- #
		if xml:	
			# -- import libraries -- #
			import openpyxl
			from openpyxl.utils import get_column_letter
			from openpyxl.styles import Font, Border, Side, Color, PatternFill, NamedStyle
			from openpyxl.chart import BarChart, Series, Reference
			import pandas as pd
			import seaborn as sns

			import matplotlib.pyplot as plt
			from ase.visualize.plot import plot_atoms
			from ase.lattice.cubic import FaceCenteredCubic
		
			# -- set plot style -- #
			sns.set_theme(style="whitegrid")

			wb = openpyxl.Workbook()

			# -- Set path -- #
			if not os.path.isdir(f'{path}/plots/geometrias/'):  	 			os.makedirs(f'{path}/plots/geometrias/')			
			if not os.path.isdir(f'{path}/plots/ORR/'): 			  			os.makedirs(f'{path}/plots/ORR/')			
			if not os.path.isdir(f'{path}/plots/metanalisis/name1'):  			os.makedirs(f'{path}/plots/metanalisis/name1')	
			if not os.path.isdir(f'{path}/plots/metanalisis/name2'):  			os.makedirs(f'{path}/plots/metanalisis/name2')	
			if not os.path.isdir(f'{path}/plots/metanalisis/name3'):  			os.makedirs(f'{path}/plots/metanalisis/name3')	
			# make magnetization folthers #
			if not os.path.isdir(f'{path}/plots/metanalisis/magnetization'):  	
				os.makedirs(f'{path}/plots/metanalisis/magnetization')
			if not os.path.isdir(f'{path}/plots/metanalisis/magnetization/System'):  
				os.makedirs(f'{path}/plots/metanalisis/magnetization/System')
			if not os.path.isdir(f'{path}/plots/metanalisis/magnetization/Functional'):  	
				os.makedirs(f'{path}/plots/metanalisis/magnetization/Functional')

			# make magnetization folthers #
			if not os.path.isdir(f'{path}/plots/metanalisis/distance'):  	
				os.makedirs(f'{path}/plots/metanalisis/distance')
			if not os.path.isdir(f'{path}/plots/metanalisis/distance/System'):  
				os.makedirs(f'{path}/plots/metanalisis/distance/System')
			if not os.path.isdir(f'{path}/plots/metanalisis/distance/Functional'):  	
				os.makedirs(f'{path}/plots/metanalisis/distance/Functional')

			# -- add styles -- # 
			order = [ get_column_letter(n) for n in range(1,50) ]
			double = Side(border_style="double", color="00008000")
			
			highlight = NamedStyle(name="highlight")
			highlight.font = Font(bold=True, size=20)
			bd = Side(style='thick', color="000000")
			highlight.border = Border(left=bd, top=bd, right=bd, bottom=bd)
			wb.add_named_style(highlight)

			redFill = PatternFill(	start_color='FFFF0000',
				                    end_color='FFFF0000',
				                    fill_type='solid')
			
			# === === ORR plot and data === === #
			if 'ORR' in feature:

				ORR_relevant = ['G1_ORR','G2_ORR','G3_ORR','G4_ORR','limiting_step_ORR','Eabs_O','Gabs_O','Eabs_OH','Gabs_OH','Eabs_OOH','Gabs_OOH','overpotencial_ORR_4e']
				ORR_matrix = []
				Magnetization_matrix = []
				Distance_matrix = []
				for set_n, (key_data, data) in enumerate(self.set.items()):
					data.ORR()
					ORR = data.reaction('ORR')
					if not os.path.exists(f"{path}/plots/ORR/{key_data}.png"):
						ORR.plot(folder=f'{path}/plots/ORR/.', name=key_data)

					if type(ORR.ORR) == dict:	 # store ORR info of all systems #
						vector_labels  = [key_data, key_data.split('_')[0], key_data.split('_')[1], '_'.join(key_data.replace('__','_').split('_')[2:]) ] 
						vector_numeric = [ORR.ORR[label] for label in ORR_relevant]

					# == M TPYP case == #
					ORR_matrix.append(vector_numeric+vector_labels)
					# == GENERAL CASE == #
					#ORR_namelist.append( [key_data, *key_data.split('_')] ) 

					for system_n, (key_system, system) in enumerate(data.system.items()):
						key_subtype = key_system.split('_')[-1] 
						try:
							if key_subtype in ['*', '*O', '*OH', '*OOH']:
								atoms_names =  np.array(system.OUTCAR.atoms_names_list)[system.CONTCAR.relevant_atoms_mask()] 
								atoms_magnetization = np.abs(system.OUTCAR.resume()['dict']['magnetization'][system.CONTCAR.relevant_atoms_mask()][:,-1])
								Magnetization_matrix += [ vector_labels+[key_subtype, atoms_names[atom_i], atom_i, atoms_magnetization[atom_i]] for atom_i, atom_name in enumerate(atoms_names)  ]	

								system.CONTCAR.atoms_names_list = system.OUTCAR.atoms_names_list
								Distance_vec = [[ vector_labels+[key_subtype, atom1, atom2, prop['distances'][pos]] for pos, atom2 in enumerate(prop['names']) ] for atom1, atom1N in system.CONTCAR.relevant_distances().items()  for num, prop in atom1N.items()  ] 
								for n in Distance_vec:
									Distance_matrix += n
						except: pass

				# -- make ORR dataframe -- #
				ORR_matrix = np.array(ORR_matrix) # store ORR info of all systems #
				ORR_df  	= pd.DataFrame(data=ORR_matrix,  columns=ORR_relevant+[f'name{i}' for i, n in enumerate(vector_labels)]) # store ORR info of all systems #
				ORR_df[ORR_relevant] = ORR_df[ORR_relevant].apply(pd.to_numeric)
				
				# -- make MARNETIZATION dataframe -- #
				Magnetization_matrix = np.array(Magnetization_matrix)
				Magnetization_df  	 = pd.DataFrame(data=Magnetization_matrix,  columns= [f'name{i}' for i, n in enumerate(vector_labels)]+['System', 'Atom', 'number', 'Magnetization'] ) # store ORR info of all systems #
				Magnetization_df['Magnetization'] = Magnetization_df['Magnetization'].apply(pd.to_numeric)

				# -- make DISTANCE dataframe -- #
				Distance_matrix = np.array(Distance_matrix)
				Distance_df  	 = pd.DataFrame(data=Distance_matrix,  columns= [f'name{i}' for i, n in enumerate(vector_labels)]+['System', 'Atom1', 'Atom2', 'Distance'] ) # store ORR info of all systems #
				Distance_df['Distance'] = Distance_df['Distance'].apply(pd.to_numeric)

				# == Magnetization Per ATOM/System/FUNCTIONAL == #  # == Magnetization Per ATOM/System/FUNCTIONAL == #  # == Magnetization Per ATOM/System/FUNCTIONAL == # 

				# Magnetization vs Functional #
				line = -2
				ws = wb.create_sheet(f'Mag_sys') # insert at the end (default)
				MAGsheet = wb.get_sheet_by_name(f'Mag_sys')
				
				ws = wb.create_sheet(f'Dist_sys') # insert at the end (default)
				DISTsheet = wb.get_sheet_by_name(f'Dist_sys')
				for name_i in pd.unique(Magnetization_df['name1']):
					line += 1

					for atom_i in pd.unique(Magnetization_df['Atom']):
						line += 1
						for i, sys_i in enumerate(pd.unique(Magnetization_df['System'])):
							# == filter data == # 
							data  = Magnetization_df[Magnetization_df['Magnetization'] < 5][Magnetization_df['Atom'] == atom_i][Magnetization_df['System'] == sys_i][Magnetization_df['name1'] == name_i]
							data2 = Distance_df[Distance_df['Distance'] < 5][Distance_df['Atom1'] == atom_i][Distance_df['System'] == sys_i][Distance_df['name1'] == name_i]

							if len(data) > 0:
								ax = sns.barplot(
								    x="name3", 
								    y="Magnetization", 
								    hue="Atom", 
								    data=data, 
								    ci="sd", 
								    edgecolor="black",
								    errcolor="black",
								    errwidth=1.5,
								    capsize = 0.1,
								    alpha=0.5
								)
								sns.stripplot(
								    x="name3", 
								    y="Magnetization", 
								    hue="Atom", 
								    data=data, dodge=True, alpha=0.6, ax=ax
								)

								handles, labels = ax.get_legend_handles_labels()
								ax.legend(handles, labels, title='Atoms', bbox_to_anchor=(1, 1.02), loc='upper left')
								ax.set_xticklabels(ax.get_xticklabels(), rotation=45)
								ax.set_title(f'Systems {sys_i} | Atom Type {atom_i} |  {name_i} ')
								fig = ax.get_figure()
								if not os.path.exists(f"{path}/plots/metanalisis/magnetization/Functional/bars_plot_{sys_i}_{atom_i}_{name_i}.png"):
									fig.savefig(f"{path}/plots/metanalisis/magnetization/Functional/bars_plot_{sys_i}_{atom_i}_{name_i}.png" , dpi=100, pad_inches=0.1, bbox_inches='tight', horizontalalignment='right') 
								plt.clf()

								try:
									img = openpyxl.drawing.image.Image(f"{path}/plots/metanalisis/magnetization/Functional/bars_plot_{sys_i}_{atom_i}_{name_i}.png")
									img.anchor = f'{get_column_letter(i*5+1)}{line*20+1}'
									img.height = 400 # insert image height in pixels as float or int (e.g. 305.5)
									img.width  = 400 # insert image width in pixels as float or int (e.g. 405.8)
									MAGsheet.add_image(img)
								except: pass								


							if len(data2) > 0:
								ax = sns.barplot(
								    x="name3", 
								    y="Distance", 
								    hue="Atom2", 
								    data=data2, 
								    ci="sd", 
								    edgecolor="black",
								    errcolor="black",
								    errwidth=1.5,
								    capsize = 0.1,
								    alpha=0.5
								)
								sns.stripplot(
								    x="name3", 
								    y="Distance", 
								    hue="Atom2", 
								    data=data2, dodge=True, alpha=0.6, ax=ax
								)

								handles, labels = ax.get_legend_handles_labels()
								ax.legend(handles, labels, title='Atoms', bbox_to_anchor=(1, 1.02), loc='upper left')
								ax.set_xticklabels(ax.get_xticklabels(), rotation=45)
								ax.set_title(f'Distance in Systems {name_i}+{sys_i} | From:{atom_i} to...  ')
								ax.set_ylim( 0.95*np.min(data2["Distance"]), 1.05*np.max(data2["Distance"]) )
								fig = ax.get_figure()
								if not os.path.exists(f"{path}/plots/metanalisis/distance/Functional/bars_plot_{sys_i}_{atom_i}_{name_i}.png"):	
									fig.savefig(f"{path}/plots/metanalisis/distance/Functional/bars_plot_{sys_i}_{atom_i}_{name_i}.png" , dpi=100, pad_inches=0.1, bbox_inches='tight', horizontalalignment='right') 
								plt.clf()

								try:
									img = openpyxl.drawing.image.Image(f"{path}/plots/metanalisis/distance/Functional/bars_plot_{sys_i}_{atom_i}_{name_i}.png")
									img.anchor = f'{get_column_letter(i*5+1)}{line*20+1}'
									img.height = 400 # insert image height in pixels as float or int (e.g. 305.5)
									img.width  = 400 # insert image width in pixels as float or int (e.g. 405.8)
									DISTsheet.add_image(img)
								except: pass								

				# == Magnetization Per ATOM/System == #	# == Magnetization Per ATOM/System == # # == Magnetization Per ATOM/System == #		
				# Magnetization vs System #	
				line = -1
				ws = wb.create_sheet(f'Mag_functional') # insert at the end (default)
				MAGsheet = wb.get_sheet_by_name(f'Mag_functional')

				ws = wb.create_sheet(f'Dist_functional') # insert at the end (default)
				DISTsheet = wb.get_sheet_by_name(f'Dist_functional')

				for atom_i in pd.unique(Magnetization_df['Atom']):
					line += 1

					for i, sys_i in enumerate(pd.unique(Magnetization_df['System'])):
						data = Magnetization_df[Magnetization_df['Magnetization'] < 5][Magnetization_df['Atom'] == atom_i][Magnetization_df['System'] == sys_i]
						data2 = Distance_df[Distance_df['Distance'] < 5][Distance_df['Atom1'] == atom_i][Distance_df['System'] == sys_i]
						
						if len(data) > 0:
							ax = sns.barplot(
							    x="name1", 
							    y="Magnetization", 
							    hue="Atom", 
							    data=data, 
							    ci="sd", 
							    edgecolor="black",
							    errcolor="black",
							    errwidth=1.5,
							    capsize = 0.1,
							    alpha=0.5
							)
							sns.stripplot(
							    x="name1", 
							    y="Magnetization", 
							    hue="Atom", 
							    data=data, dodge=True, alpha=0.6, ax=ax
							)

							handles, labels = ax.get_legend_handles_labels()
							ax.legend(handles, labels, title='Atoms', bbox_to_anchor=(1, 1.02), loc='upper left')
							ax.set_xticklabels(ax.get_xticklabels(), rotation=45)
							ax.set_title(f'Atom Type {atom_i} |  {name_i} ')
							fig = ax.get_figure()
							if not os.path.exists(f"{path}/plots/metanalisis/magnetization/System/bars_plot_{sys_i}_{atom_i}.png"):	
								fig.savefig(f"{path}/plots/metanalisis/magnetization/System/bars_plot_{sys_i}_{atom_i}.png" , dpi=100, pad_inches=0.1, bbox_inches='tight', horizontalalignment='right') 
							plt.clf()

							try:
								img = openpyxl.drawing.image.Image(f"{path}/plots/metanalisis/magnetization/System/bars_plot_{sys_i}_{atom_i}.png")
								img.anchor = f'{get_column_letter(i*5+1)}{line*20+1}'
								img.height = 400 # insert image height in pixels as float or int (e.g. 305.5)
								img.width  = 400 # insert image width in pixels as float or int (e.g. 405.8)
								MAGsheet.add_image(img)
							except: pass	

						if len(data) > 0:
							ax = sns.barplot(
							    x="name1", 
							    y="Distance", 
							    hue="Atom2", 
							    data=data2, 
							    ci="sd", 
							    edgecolor="black",
							    errcolor="black",
							    errwidth=1.5,
							    capsize = 0.1,
							    alpha=0.5
							)
							sns.stripplot(
							    x="name1", 
							    y="Distance", 
							    hue="Atom2", 
							    data=data2, dodge=True, alpha=0.6, ax=ax
							)

							handles, labels = ax.get_legend_handles_labels()
							ax.legend(handles, labels, title='Atoms', bbox_to_anchor=(1, 1.02), loc='upper left')
							ax.set_xticklabels(ax.get_xticklabels(), rotation=45)
							ax.set_title(f'Distance in {sys_i} | From: {atom_i} to...  ')
							ax.set_ylim( 0.95*np.min(data2["Distance"]), 1.05*np.max(data2["Distance"]) )
							fig = ax.get_figure()
							if not os.path.exists(f"{path}/plots/metanalisis/distance/System/bars_plot_{sys_i}_{atom_i}.png"):	
								fig.savefig(f"{path}/plots/metanalisis/distance/System/bars_plot_{sys_i}_{atom_i}.png" , dpi=100, pad_inches=0.1, bbox_inches='tight', horizontalalignment='right') 
							plt.clf()


							try:
								img = openpyxl.drawing.image.Image(f"{path}/plots/metanalisis/distance/System/bars_plot_{sys_i}_{atom_i}.png")
								img.anchor = f'{get_column_letter(i*5+1)}{line*20+1}'
								img.height = 400 # insert image height in pixels as float or int (e.g. 305.5)
								img.width  = 400 # insert image width in pixels as float or int (e.g. 405.8)
								DISTsheet.add_image(img)
							except: pass	

				# === META analisis ORR === # # === META analisis ORR === # # === META analisis ORR === # # === META analisis ORR === # # === META analisis ORR === #
				for orr_r in ORR_relevant:		
					line = 0
					ws = wb.create_sheet(orr_r) # insert at the end (default)
					ORRsheet = wb.get_sheet_by_name(orr_r)

					# ============= PLOT name1 ( system name ) ============= #
					if not os.path.isdir(f'{path}/plots/metanalisis/name1/{orr_r}'):  os.makedirs(f'{path}/plots/metanalisis/name1/{orr_r}')
					for name1 in pd.unique(ORR_df['name1']):
						if not os.path.exists(f"{path}/plots/metanalisis/name1/{orr_r}/bars_plot_{name1}.png"):
							ORR_df2 = ORR_df[ ORR_df['name1']== name1 ]
							ax = sns.barplot(x="name3", y=orr_r, hue='name2', data=ORR_df2)
							ax.set_title(name1)
							fig = ax.get_figure()
							for tick in ax.xaxis.get_major_ticks():
								tick.label.set_fontsize(10) 
							ax.set_xticklabels(ax.get_xticklabels(), rotation=45)
							fig.savefig(f"{path}/plots/metanalisis/name1/{orr_r}/bars_plot_{name1}.png", dpi=100, pad_inches=0.1, bbox_inches='tight', horizontalalignment='left') 
							plt.clf()
						try:
							img = openpyxl.drawing.image.Image(f"{path}/plots/metanalisis/name1/{orr_r}/bars_plot_{name1}.png")
							img.anchor = f'{get_column_letter((line%3)*5+1 )}{int(line/3)*20+1}'
							img.height = 400 # insert image height in pixels as float or int (e.g. 305.5)
							img.width  = 400 # insert image width in pixels as float or int (e.g. 405.8)
							ORRsheet.add_image(img)
						except: pass
						line += 1

					# ============= PLOT name3 ( system functional ) ============= #
					line += 10
					if not os.path.isdir(f'{path}/plots/metanalisis/name3/{orr_r}'):  os.makedirs(f'{path}/plots/metanalisis/name3/{orr_r}')
					for name3 in pd.unique(ORR_df['name3']):
						if not os.path.exists(f"{path}/plots/metanalisis/name3/{orr_r}/bars_plot_{name3}.png"):
							ORR_df2 = ORR_df[ ORR_df['name3']== name3 ]
							ax = sns.barplot(x="name1", y=orr_r, hue='name2', data=ORR_df2)
							ax.set_title(name3)
							fig = ax.get_figure()
							for tick in ax.xaxis.get_major_ticks():
								tick.label.set_fontsize(10) 
							ax.set_xticklabels(ax.get_xticklabels(), rotation=45)
							fig.savefig(f"{path}/plots/metanalisis/name3/{orr_r}/bars_plot_{name3}.png" , dpi=100, pad_inches=0.1, bbox_inches='tight', horizontalalignment='right') 
							plt.clf()

						try:
							img = openpyxl.drawing.image.Image(f"{path}/plots/metanalisis/name3/{orr_r}/bars_plot_{name3}.png")
							img.anchor = f'{get_column_letter((line%3)*5+1 )}{int(line/3)*20+1}'
							img.height = 400 # insert image height in pixels as float or int (e.g. 305.5)
							img.width  = 400 # insert image width in pixels as float or int (e.g. 405.8)
							ORRsheet.add_image(img)
						except: pass
						line += 1

				# === ORR TAB === #
				ws = wb.create_sheet('ORR') # insert at the end (default)
				ORRsheet = wb.get_sheet_by_name('ORR')
				ORRsheet.column_dimensions['A'].width = 30
				for i, label in enumerate(['Sys name','dG1','dG2','dG3','dG4','Lim step',
											'Eabs_O','Gabs_O','Eabs_OH','Gabs_OH','Eabs_OOH','Gabs_OOH','OP']):
					ORRsheet['{}{}'.format(get_column_letter(i+1), 2)].value = label
					ORRsheet['{}{}'.format(get_column_letter(i+1), 2)].font = Font(name="Tahoma", size=13, color="00339966", bold=True)
					ORRsheet['{}{}'.format(get_column_letter(i+1), 2)].border = Border(top=double, bottom=double)

			for set_n, (key_data, data) in enumerate(self.set.items()):
				if not os.path.isdir(f'{path}/plots/systemas/{key_data}'):  	  os.makedirs(f'{path}/plots/systemas/{key_data}')	

				# == make new sheet == (with an acceptable name)
				sheet_name = key_data.replace('*','x') 
				sheet_name = sheet_name.replace('[','L') 
				sheet_name = sheet_name.replace(']','L') 

				ws = wb.create_sheet(sheet_name) # insert at the end (default)
				sheet = wb.get_sheet_by_name(sheet_name)

				# == set some sheer variables ==
				sheet.column_dimensions['A'].width = 30
				head_space = 37 # space from first line
				line = 1+head_space 
				for system_n, (key_system, system) in enumerate(data.system.items()):
					# ======================= PLOT ATOMS ======================= #
					if system.CONTCAR != None:	
						try:
							#if not os.path.exists(f'{path}/plots/systemas/{key_data}/{key_system}.png'):
								#system.CONTCAR.plot(save=True, path=f'{path}/plots/systemas/{key_data}/{key_system}.png')
							img = openpyxl.drawing.image.Image(f'{path}/plots/systemas/{key_data}/{key_system}.png')
							img.anchor = f'A{line}'
							sheet.add_image(img)
						except: pass

					# ======================= PRINT summary ======================= #
					# = System label = #
					sheet.row_dimensions[line-1].height = 40
					sheet[f'A{line-1}'].value = key_system
					sheet[f'A{line-1}'].font   = Font(name="Tahoma", size=20, color="00339966", bold=True)
					sheet[f'A{line-1}'].border = Border( bottom=double )

					sheet[f'D{line-1}'].value = key_system
					sheet[f'D{line-1}'].font   = Font(name="Tahoma", size=20, color="00339966", bold=True)
					sheet[f'D{line-1}'].border = Border( bottom=double )

					for load in system.loaded:
						# = Files name = #
						sheet[f'D{line}'].value  = load
						sheet[f'D{line}'].font   = Font(name="Tahoma", size=13, color="00339966", bold=True)
						sheet[f'D{line}'].border = Border( right=double )			
						
						# = Files summary = #
						try:
							if 	 load == 'POSCAR': system_resume = system.POSCAR.resume()
							elif load == 'CONTCAR': system_resume = system.CONTCAR.resume()
							elif load == 'OSZICAR': system_resume = system.OSZICAR.resume()
							elif load == 'OUTCAR': system_resume = system.OUTCAR.resume()
							else: system_resume= {'dict':{}}
							#print( system.CONTCAR.relevant_distances() )

							for i_load, (key_load,	 value_load) in enumerate( system_resume['dict'].items() ):
								sheet.cell(row=line, column=i_load+5).value = '{} {}'.format(str(key_load), str(value_load))
								sheet.column_dimensions[get_column_letter(i_load+2)].width = 15
						except:	print('WARNING :: Set.summary() :: can NOT plot summarise POSCAR/CONTCAR/OSZICAR/OUTCAR')
						line += 1
					line += 20

				# === === === ORR === === === # # === === === ORR === === === # # === === === ORR === === === #
				if 'ORR' in feature:
					link = './summary.xlsx'+f'#{sheet_name}!A1'

					ORRsheet[f'A{set_n+3}'].hyperlink = link
					ORRsheet[f'A{set_n+3}'].value = '#'+sheet_name
					ORRsheet[f'A{set_n+3}'].style = "Hyperlink"

					# == PLOT == #
					try:
						img = openpyxl.drawing.image.Image(f'./{path}/plots/ORR/{key_data}.png' )
						img.anchor = 'A1'
						sheet.add_image(img)
					except: print(f'Can not find ./{path}/plots/ORR/{key_data}.png')

					try:
						ORR = data.ORR()
						for i, label in enumerate(ORR_relevant):
							if type(ORR.ORR) != type(None):
								ORRsheet[f'{get_column_letter(i+2)}{set_n+3}'].value = ORR.ORR[label]
						
						sheet['B1'].style = 'highlight'
					except:	print('WARNING :: Set.summary() :: can NOT plot summarise xlsx ORR')

			wb.save('{}/summary.xlsx'.format(path) )

		return True

	def extract_features(self, feature=None, v=True, filters='None'):
		# extract specific features 
		# ----------------------------------------------------------------------------------------

		# ----------------------------------------------------------------------------------------
		# feature 	::	dict 	:: feature to extract 
		# v 		::  BOOL 	:: vervosity eg: True
		# filter 	::  STR 	:: filter on features space

		'''
		feature = { 'ORR' : ['overpotencial_ORR_4e']
					'PDOS': {	'config' : {'start':-1.0, 'end':1.0, 'point':500}
								'atoms'  : {'name':['Co', 'Fe']}} ,  or {'closest':'_*O'}
								'orbital': ['d']  } 
					'magnetization'	: {'atoms'  : {'name':['Co', 'Fe']}} 
					'charge'		: {'atoms'  : {'name':['Co', 'Fe']}}
					}	
		'''
		# -- set data storage -- #
		feature_ext = {} # here we extract all data
		if v: print('*'*10 + '\nExtracting features\n'+'*'*10) # verbosity 

		# *********** ITERATION data *********** #
		for set_n, (key_data, data) in enumerate(self.set.items()):
			key_data = key_data.split('\n')[-1]
			if v: print(' [{1}] reading system :: {0} :: '.format(str(key_data), str(set_n)) ) # verbosity 

			# ------- ORR ------- #
			if 'ORR' in feature:
				try:
					ORR = data.ORR()
					if not key_data in feature_ext: feature_ext[key_data] = {}
					feature_ext[key_data]['ORR'] = [ ORR.ORR[orr_feature] for orr_feature in feature['ORR'] ]
				except:
					feature_ext[key_data]['ORR'] = [None]

				if v: 
					try:
						print(' [{1}] :: {0} :: evaluating ORR parameters.'.format(str(key_data), str(set_n)) ) # verbosity 				
						for orr_feature in feature['ORR']:	print(' [{3}] :: {0} :: ORR :: {1} {2} '.format(str(key_data), str(orr_feature), str(ORR.ORR[orr_feature]), str(set_n), ) ) # verbosity xxx
					except:
						print('WARNING :: Set.extract_features() :: can NOT print OOR summary')

			# *********** ITERATION system *********** #
			for system_n, (key_system, system) in enumerate(data.system.items()):
				key_system = key_system.split('\n')[-1]
				if v: print(' [{2}.{3}] :: {0} :: {1} ::'.format(str(key_data), str(key_system), str(set_n), str(system_n) )) # verbosity 
				
				# ------- DOSCAR/ ------- #
				if system.DOSCAR != None and 'PDOS' in key_system and 'PDOS' in feature:
					if v: print(' [{2}.{3}] :: {0} :: {1} :: PDOS'.format(str(key_data), str(key_system), str(set_n), str(system_n)  )) # verbosity 
					DOSCAR_feature = []

					# atom_positions #
					if 'closest' in feature['PDOS']['atoms'] :
						atom_positions = []
						for (key_system1, system1) in data.system.items():
							keyword = feature['PDOS']['atoms']['closest']
							if keyword == key_system1[-int(len(keyword)):] and type(system1.POSCAR.atoms_names) != type(None): # particular de mi sistema
								atom_positions = [ system1.POSCAR.closest('O')[1] ]
					if 'name' in feature['PDOS']['atoms']:
						atom_positions = []
						for a in feature['PDOS']['atoms']['name']:
							atom_positions += system.POSCAR.get_postion(a)
					# ************** #
					
					for orbital in feature['PDOS']['orbital']:
						config = feature['PDOS']['config']

						for ap in atom_positions:
							DOSCAR_feature.append( system.DOSCAR.cut( 	ion=ap, orbital=orbital, 
																start= 	config['start'], 
																end=	config['end'],
																point=	config['point'], ) )

				else: DOSCAR_feature = None
				# ------- /DOSCAR ------- #

				# ------- relevant_distances  ------- #
				if system.CONTCAR != None and 'relevant_distances' in feature:
					system.CONTCAR.relevant_distances(relevance=2, relevance_distance=2.5, save=True, v=v)

				# ------- magnetization/  ------- #
				if system.POSCAR != None and system.OUTCAR != None and 'magnetization' in feature:
				
					# atom_positions #
					if 'closest' in feature['magnetization']['atoms'] :
						atom_positions = []
						for (key_system1, system1) in data.system.items():
							keyword = feature['magnetization']['atoms']['closest']
							if keyword == key_system1[-int(len(keyword)):] and type(system1.POSCAR.atoms_names) != type(None): # particular de mi sistema
								atom_positions = [ system1.POSCAR.closest('O')[1] ]

					elif 'name' in feature['magnetization']['atoms'] :
						atom_positions = []
						for a in feature['magnetization']['atoms']['name']:
							atom_positions += system.POSCAR.get_postion(a)
					magnetization_feature = system.OUTCAR.get_magnetization( atom_positions )
					if v: print(' [{2}.{3}] :: {0} :: {1} :: magnetization {4}'.format(str(key_data), str(key_system), str(set_n), str(system_n), magnetization_feature)) # verbosity 

				else: magnetization_feature = None
				# ------- /magnetization  ------- #

				# ------- charge/  ------- #
				if system.OUTCAR != None and 'charge' in feature:
				
					# atom_positions #
					if 'closest' in feature['charge']['atoms'] :
						atom_positions = []
						for (key_system1, system1) in data.system.items():
							keyword = feature['charge']['atoms']['closest']
							if keyword == key_system1[-int(len(keyword)):] and type(system1.POSCAR.atoms_names) != type(None): # particular de mi sistema
								atom_positions = [ system1.POSCAR.closest('O')[1] ]
					elif 'name' in feature['charge']['atoms'] :
						atom_positions = []
						for a in feature['charge']['atoms']['name']:
							atom_positions += system.POSCAR.get_postion(a)
					charge_feature = system.OUTCAR.get_charge( atom_positions )
					if v: print(' [{2}.{3}] :: {0} :: {1} :: charge {4}'.format(str(key_data), str(key_system), str(set_n), str(system_n), charge_feature)) # verbosity 
				# ------- /charge  ------- #

				else: charge_feature = None

				if filters == 'AND':
					# -- AND filter -- #
					check_list = [ 	charge_feature if 'charge' in feature else None,
									magnetization_feature if 'magnetization' in feature else None,
									DOSCAR_feature if 'PDOS' in feature else None,
									ORR.ORR if 'ORR' in feature else None,
									 ]

					if  sum([ 1 if type(n) != type(None) else 0 for n in check_list ]) == len(check_list):
						if not key_data in feature_ext: feature_ext[key_data] = {}
						if not key_system in feature_ext[key_data]: feature_ext[key_data][key_system] = {}

						if 'charge' in feature_ext[key_data]:			feature_ext[key_data][key_system]['charge'] = charge_feature

						if 'magnetization' in feature_ext[key_data]:	feature_ext[key_data][key_system]['magnetization'] = magnetization_feature 

						if 'PDOS' in feature_ext[key_data]:				feature_ext[key_data][key_system]['PDOS'] = np.concatenate(DOSCAR_feature)

				else:
					# -- with out any filter -- #
					if not key_data in feature_ext: feature_ext[key_data] = {}
					if not key_system in feature_ext[key_data]: feature_ext[key_data][key_system] = {}

					if 'charge' in feature:				feature_ext[key_data][key_system]['charge'] = charge_feature

					if 'magnetization' in feature: 		feature_ext[key_data][key_system]['magnetization'] = magnetization_feature

					if 'PDOS' in feature:				feature_ext[key_data][key_system]['PDOS'] = DOSCAR_feature

		self.feature_ext = feature_ext
		
		return feature_ext 

	def feature_dic2array(self, dictionary, ):
		# transform dictionary to the corresponding arrays

		dictionary = dictionary if type(dictionary) == dict else self.feature_ext
		#dic_array = { key_data:[] for (key_data, data) in dictionary.items() }
		dic_array = {}
		list_name = [ key_label for (key_label, value_label) in dictionary.items() ]

		for (key_label, value_label) in dictionary.items():
			for (key_data, value_data) in value_label.items():
				for value in value_data:
					if key_data in dic_array:
						dic_array[key_data] = np.concatenate( (dic_array[key_data], np.array(value)) )
					else:
						dic_array[key_data] = np.array(value)

		for (key_data, data) in dic_array.items():
			dic_array[key_data] = np.array(data)

		return dic_array, list_name

	def get_dataset(self, load_path=None, save_path=None, Xfeatures=None, Yfeatures=None):
		# this function generates a complete dataset 
		# 	X 	[samples, features] 	
		# 	Y 	[samples, features]
		# 	X --> Y

		if type(load_path) == str: 	self.load_data( filename=load_path)
		Xfeatures = {'PDOS': {	'config' : {'start':-5.0, 'end':3.0, 'point':500},
										'atoms'  : {'name':['Fe', 'Co', 'Cu', 'Mg']},  
										'orbital': [9,10,11,12,13,14,15,16,17,18,]  },
					'magnetization'	: {'atoms'  : {'name':['Fe', 'Co', 'Cu', 'Mg']},},
					'charge'		: {'atoms'  : {'name':['Fe', 'Co', 'Cu', 'Mg']},},} if type(Xfeatures) != dict else Xfeatures 
		
		
		Yfeatures = { 		
							'ORR' : [	'overpotencial_ORR_4e', 'Gabs_OOH', 'Eabs_OOH', 'Gabs_OH', 
										'Eabs_OH', 'Gabs_O', 'Eabs_O', 'G1_ORR', 'G2_ORR', 'G3_ORR', 'G4_ORR', 'limiting_step_ORR'],
							} if type(Yfeatures) != dict else Yfeatures 

		features = { key:value if not key in Xfeatures else Xfeatures[key] for (key,value) in list(Xfeatures.items())+list(Yfeatures.items()) }
		features_ext = self.extract_features(feature=features, v=False )
		
		NNx, NNy, names = [], [], []

		for key_data, value_data in features_ext.items():
			for key_system, value_system in value_data.items():

				if 'PDOS' in key_system and type(features_ext[key_data]['ORR']) != type(None) and type(value_system['PDOS']) != type(None) and len(features['ORR']) == len(features_ext[key_data]['ORR']):
					NN_in = []
					if 'PDOS' in Xfeatures: 			NN_in.append( np.array(features_ext[key_data][key_system]['PDOS']).flatten()   )
					if 'magnetization' in Xfeatures: 	NN_in.append( features_ext[key_data][key_system]['magnetization'][0] )
					if 'charge' in Xfeatures: 			NN_in.append( features_ext[key_data][key_system]['charge'][0]   )

					NNx.append( np.concatenate(NN_in) )
					NNy.append( features_ext[key_data]['ORR'] )
					names.append( key_data )

		if type(save_path) == str:
			np.savetxt('{}/NNx.dat'.format(save_path), NNx)
			np.savetxt('{}/NNy.dat'.format(save_path), NNy)
			np.savetxt('{}/names.dat'.format(save_path), names, fmt="%s")

		return np.array(NNx), np.array(NNy), np.array(names)

	# **************************************************************************************************************************************************************
	# **************************************************************************************************************************************************************
	# **************************************************************************************************************************************************************
	def embedding(self, action, **kwargs):
		
		def save_emb(path, ff_X, ff_Y):
			if not os.path.isdir(path):  os.makedirs(path)

			for atom, ff_x in ff_X.items():
				if not os.path.isdir(f'{path}/{atom}'):  os.makedirs(f'{path}/{atom}')
				with open(f'{path}/{atom}/ff_X.dat', "ab") as file:
					np.savetxt(file, ff_x )

			for atom, ff_y in ff_Y.items():
				with open(f'{path}/{atom}/ff_Y.dat', "ab") as file:
					np.savetxt(file, ff_y )

			return None

		def load_emb(path, v=True):
			print(path)
			Xembedding_dict = {}
			Yembedding_dict = {}
			if v: print('Loading embedding ...')
			
			for i, subfolder in enumerate(os.listdir(path)):
				t1 = time.time()
				Xembedding_dict[str(subfolder)] = np.loadtxt(f'{path}/{subfolder}/ff_X.dat' )
				Yembedding_dict[str(subfolder)] = np.loadtxt(f'{path}/{subfolder}/ff_Y.dat' )
				if v: print(f'\t ({i}) Load atom: {subfolder} \t Samples: {Xembedding_dict[str(subfolder)].shape[0]} \t Dimention {Xembedding_dict[str(subfolder)].shape[1]} \t ({time.time()-t1}s) ')

			return Xembedding_dict, Yembedding_dict

		def make_emb(embedding_parameters=None, path=None, v=True, save=True):
			# === Get data from OUTCAR === #
			embedding_parameters = {'partition':400, 'min distance':0.5, 'max distance':7.0, 'range distance':0.001} if embedding_parameters == None else embedding_parameters
					
			for dataset_n, (key_dataset, dataset) in enumerate(self.set.items()):
				if v: print(' Getting embeding from  {}'.format(key_dataset) )
				result_list = dataset.get_embeddings(embedding_parameters=embedding_parameters, path=path, 
													processing_pool={'multiprocessing':True, 'cores':'check'}, save=False)

				if save:
					for result in result_list:	
						save_emb(path, result['ff_X'], result['ff_Y'])

			return load_emb(path, v=v)

		if action == 'load': return load_emb(**kwargs)
		if action == 'make': return make_emb(**kwargs)
		if action == 'save': return save_emb(**kwargs)


	def forcefield(self, embedding='make', embedding_parameters=None, 
						 forcefield='make', MLPR=None, 
						 path=None,  save=True, v=True):
		'''
		dataset_path = '/home/akaris/Documents/code/VASP/v3.5/files/dataset/force_trainning/dataset_ff.pkl'
		ff_path = '/home/akaris/Documents/code/VASP/v3.5/files/dataset/force_trainning/force_field/02'
		dataset = Set()
		dataset.load_data( filename=dataset_path )
		MLPR = dataset.forcefield( path=ff_path, embedding='None', forcefield='load' )
		'''

		def load_forcefield( path=None, v=True, save=True):
			MLPR = {} 

			for i, subfolder in enumerate(os.listdir(path)):
				t1 = time.time()

				try:
					with open(f'{path}/{subfolder}/regrNN', 'rb') as file:
						MLPR[str(subfolder)] = pickle.load(file)
						if v: print(f'Reading atom: {subfolder} \t time {time.time()-t1}')

				except:
					if v: print(f'Can not find file: *regrNN* in path {path}/subfolder/')

			return MLPR

		def save_forcefield( path=None , regr=None):
			with open(f'{path}/regrNN', 'wb') as file:
				pickle.dump(regr, file)
		
		def train_forcefield( ff_X, ff_Y, MLPR=None, v=True, save=True ):
			MLPR = {} if type(MLPR) != dict else MLPR
			for atom, data in ff_X.items():
				# ========== Filter data ========== #
				ffx_atom_train, ffy_atom_train = ff_X[str(atom)], ff_Y[str(atom)]
				filters = np.sum(np.abs( ffy_atom_train ),axis=1) < 0.5

				ffx_atom_train = ffx_atom_train[filters,:]
				ffy_atom_train = ffy_atom_train[filters,:]

				# ========== Train model NN ========== #
				if type(MLPR) == dict and str(atom) in MLPR:
					regr_atom = MLPR[atom].partial_fit(ffx_atom_train, ffy_atom_train)
				else:
					regr_atom = MLPRegressor(random_state=3, activation='tanh', #early_stopping=True, validation_fraction=0.10,
									hidden_layer_sizes=(30, 30, 30, 30),max_iter=1000, #learning_rate='adaptive', warm_start=True,
									#hidden_layer_sizes=(  400, 540, 540, 400,  ),max_iter=5000,
									momentum=0.7, verbose=True, n_iter_no_change=180, 
									early_stopping=True, validation_fraction=0.1).fit(ffx_atom_train, ffy_atom_train)

				MLPR[str(atom)] = regr_atom

				# ========== SAVE model NN ========== #
				if not os.path.isdir(f'{path}/{atom}'):  os.makedirs(f'{path}/{atom}')
				if save: save_forcefield( path=f'{path}/{atom}' , regr=regr_atom)

			return MLPR

		def validate_forcefield( path, MLPR, ff_X, ff_Y, v=True, save=True ):
			for key, value in ff_X.items():
				filters = np.sum(np.abs( ff_Y[key] ),axis=1) < 1
				value = value[filters,:]
				ff_Y[key] = ff_Y[key][filters,:]

				ff_y = MLPR[key].predict( value )
				plt.plot( ff_Y[key][:,0], ff_y[:,0], 'o', c='r', ) 
				plt.plot( ff_Y[key][:,1], ff_y[:,1], 'o', c='g', ) 
				plt.plot( ff_Y[key][:,2], ff_y[:,2], 'o', c='b', ) 
				plt.title(f'{key}')
				plt.show()

		# === Embedding === #
		if   embedding == 'make': ff_X, ff_Y = self.embedding('make', embedding_parameters=embedding_parameters, path=path, v=v, save=save)
		elif embedding == 'load': ff_X, ff_Y = self.embedding('load', path=path, v=v, )
		else: ff_X, ff_Y = None, None

		# === Multi-layer Perceptron regressor === #
		if   forcefield == 'make': MLPR = train_forcefield( ff_X, ff_Y, MLPR, v=v, save=save )
		elif forcefield == 'load': MLPR = load_forcefield( path, v=v, save=save )
		elif forcefield == 'validate': 
			MLPR=load_forcefield( path, v=v, save=save )
			validate_forcefield( path, MLPR, ff_X, ff_Y, v=v, save=save )
		else: MLPR = {}

		if save: self.ff_MLPR = MLPR

		return MLPR

def main(argv):
	# === organize arg === #
	inputfile  = argv['input']
	outputfile = argv['output']
	outputfile = inputfile if type(outputfile) == type(None) else outputfile
	task 	   = argv['task']

	# === Make data holder === #
	dataset = Set()

	if task == 'read':
		data_dict = __import__(inputfile)
		dataset.read_data( dataset=data_dict.Set, verbosity=True)
		filename = f'{outputfile}.pkl'
		filehandler = open(filename, 'wb')
		pickle.dump(dataset, filehandler)

	if task == 'xml':
		dataset.load_data( filename=inputfile, verbosity=True)
		dataset.summary(xml=True, path=outputfile)
		
	print(f'Input  >> {inputfile} ')
	print(f'OUTPUT >> {outputfile}')

if __name__ == "__main__":
	parser = argparse.ArgumentParser()
	
	parser.add_argument('-t','--task', help="task to accomplish \n   read  :  Read files from dataset  \n    summarise  :  resume data from dataset ",
	                    type=str, default='read', required=True)
	
	parser.add_argument('-i','--input', help="path to inputfile",
	                    type=str, default='', required=True)

	parser.add_argument('-o','--output', help="name of data output file",
	                    type=str, default=None, required=False)

	args = vars(parser.parse_args())
	main(args)

'''
'''

'''
dataset_path = '/home/akaris/Documents/code/VASP/v3.5/files/dataset/force_trainning/dataset_ff.pkl'
ff_path = '/home/akaris/Documents/code/VASP/v3.5/files/dataset/force_trainning/force_field/03'
dataset = Set()
dataset.load_data( filename=dataset_path )
MLPR = dataset.forcefield( path=ff_path, embedding='load', forcefield='make', embedding_parameters={'partition':5, 'min distance':0.5, 'max distance':7.0, 'range distance':0.001} )
MLPR = dataset.forcefield( path=ff_path, embedding='load', forcefield='validate', embedding_parameters={'partition':5, 'min distance':0.5, 'max distance':7.0, 'range distance':0.001} )

OC = OUTCAR.OUTCAR()
OC.load('/home/akaris/Documents/code/VASP/v3.5/files/OUTCAR/FeTPyP/OUTCAR')

path = '/home/akaris/Documents/code/VASP/v3.5/files/dataset/force_trainning/dynamics/path10'
position = OC.total_force[5,:,:3]
velocidad = np.zeros_like(position)
force = np.zeros_like(position)
dt = 0.01
for n in range(800):
	forces = OC.get_NNforces(position=position, model=MLPR, embedding_parameters={'partition':5, 'min distance':0.5, 'max distance':7.0, 'range distance':0.001}, save=True)
	velocidad[:,:] -= forces[:,:]*dt*0.5 
	position  += velocidad*dt*0.5 
	OC.save_trajectory_step(atoms_position=position, file_name=path, append_data=True, traj_step=None, traj_time=None,save=True)
	print(n)

plt.show()
'''

'''
# eg. Summary  *************************************************** ***************************************************
path = 	'../files/dataset/MePC/M1PC.pkl'
dataset = Set()
dataset.load_data( filename=path )
dataset.plot_ORR_summary(path='', save=False, v=True)
'''

'''
dataset.summary(xml=True, path='ORR')
'''


'''
# eg. NN data extraction #  *************************************************** ***************************************************
path = 	'../files/dataset/Metales/MeTPyP/MeTPyP.pkl'
dataset = Set()
dataset.load_data( filename=path )

NNx, NNy = dataset.get_dataset() # [samples, features] 	[samples, features]
'''

# eg. NN data extraction and analisys # *************************************************** ***************************************************
'''
# ==== LOAD data and store training matrix ==== #
path = 	'../files/dataset/Metales/MePC/MePC.pkl'
path = 	'../files/dataset/Metales/MeTPyP/MeTPyP.pkl'
dataset = Set()
#dataset.load_data( filename=path )
#NNx, NNy, names = dataset.get_dataset( save_path='/home/akaris/Documents/code/VASP/v3.5/files/dataset/Metales/MeTPyP' ) # [samples, features] 	[samples, features]

# ==== LOAD training matrix ==== #
NNx_PC = np.loadtxt('/home/akaris/Documents/code/VASP/v3.5/files/dataset/Metales/MePC/NNx.dat')
NNy_PC = np.loadtxt('/home/akaris/Documents/code/VASP/v3.5/files/dataset/Metales/MePC/NNy.dat')
names_PC = np.loadtxt('/home/akaris/Documents/code/VASP/v3.5/files/dataset/Metales/MePC/names.dat', dtype=str)

NNx_TPyP = np.loadtxt('/home/akaris/Documents/code/VASP/v3.5/files/dataset/Metales/MeTPyP/NNx.dat')
NNy_TPyP = np.loadtxt('/home/akaris/Documents/code/VASP/v3.5/files/dataset/Metales/MeTPyP/NNy.dat')
names_TPyP = np.loadtxt('/home/akaris/Documents/code/VASP/v3.5/files/dataset/Metales/MeTPyP/names.dat', dtype=str)

NNx = np.concatenate( (NNx_PC, NNx_TPyP) )
NNy = np.concatenate( (NNy_PC, NNy_TPyP) )
names = np.concatenate( (names_PC, names_TPyP) )

#NNx = StandardScaler().fit_transform( NNx )
#NNy = StandardScaler().fit_transform( NNy )

features = [	'overpotencial_ORR_4e', 'Gabs_OOH', 'Eabs_OOH', 'Gabs_OH', 'Eabs_OH', 'Gabs_O', 'Eabs_O', 
				'G1_ORR', 'G2_ORR', 'G3_ORR', 'G4_ORR', 'limiting_step_ORR']

# ==== Split trainning/test data ==== #
train_index = [ True if  not 'Ru' in name and not 'Mg' in name and not 'Pt' in name  and not 'Co' in name else False for name in names]
test_index = [ True if 'PtPc' in name  or 'CoPc' in name  else False for name in names]
X_train, X_test = NNx[train_index], NNx[test_index] 
y_train, y_test = NNy[train_index], NNy[test_index]
names_train, names_test = names[train_index], names[test_index], 

# ==== Make embedding vector ==== #
embedding_train = np.array([ [ 1 if m+1 == n else 0 for m in range(4)] for n in y_train[:,-1] ])
embedding_test  = np.array([ [ 1 if m+1 == n else 0 for m in range(4)] for n in y_test[:,-1] ])

# ==== Model trainning ==== # ['identity', 'logistic', 'relu', 'softmax', 'tanh'].
regr = MLPRegressor(random_state=50, activation='relu', #early_stopping=True, validation_fraction=0.10,
					hidden_layer_sizes=(100, 20, 20, 20, 100, ),max_iter=10000, 
					#early_stopping=True, validation_fraction=0.1, solver='adam', 
					#learning_rate='adaptive', warm_start=True, solver=’sgd’ or ‘adam’.
					momentum=0.9, verbose=True, n_iter_no_change=20, ).fit(X_train, y_train[:,0])

# ==== Results analisys ==== #
NNp_test =  regr.predict(X_test) # [samples, features]
plt.figure(1), plt.plot([0,1], [0,1], '--')
plt.plot(NNp_test[:], y_test[:, 0], 'o')
for i, yt in enumerate(y_test[:, feature_test]):		
	plt.text(NNp_test[i]*2, y_test[i, feature_test]*2, names_test[i], bbox=dict(facecolor='red', alpha=0.04))
plt.show()
'''









